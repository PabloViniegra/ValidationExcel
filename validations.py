import re
import datetime
import pandas as pd
from openpyxl import load_workbook

pathToFiles = '../FTP/'


class PremiamosTuConfianzaDentalPre:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Analizando " + file + " ....")
                df = pd.read_excel(pathToFiles + file)
                line_number = 2
                for index, row in df.iterrows():
                    # validación de idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS")
                        workbook = load_workbook(filename=pathToFiles + file)
                        sheet = workbook.active
                        sheet["C" + str(line_number)] = "CAS"
                        workbook.save(pathToFiles + file)
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS" + "\n")
                    # validación si el numero de tarjeta o el nif vienen vacíos
                    if pd.isna(row['N_TARJETA']) and pd.isna(row['NIF']):
                        print("La fila " + str(line_number) + " no tiene número de tarjeta ni NIF en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene número de tarjeta ni NIF en " + file + "\n")
                        checkToReturn = False
                    # validación de la vía de impacto
                    if pd.isna(row['CORREO_CLIENTE']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL_SMS']) and pd.isna(row['TELEFONO_MOVIL_SMS_2']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS']), re.M) or re.fullmatch(
                                    "(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS_2']), re.M):
                                print(
                                    "La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn


class VentajasBasico:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Analizando " + file + " ....")
                df = pd.read_excel(pathToFiles + file)
                workbook = load_workbook(filename=pathToFiles + file)
                sheet = workbook.active
                line_number = 2
                # validacion de si en target corresponde JOV o GRAL
                if '_JOV_' in file:
                    targetToAdd = 'JOVEN'
                    print("En el archivo " + file + " el target debe ser " + targetToAdd)
                    with open('../OK.txt', 'a') as f:
                        f.write("En el archivo " + file + " el target debe ser " + targetToAdd + "\n")
                else:
                    targetToAdd = 'GRAL'
                    print("En el archivo " + file + " el target debe ser " + targetToAdd)
                    with open('../OK.txt', 'a') as f:
                        f.write("En el archivo " + file + " el target debe ser " + targetToAdd + "\n")
                for index, row in df.iterrows():
                    # validación del idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS")
                        sheet["C" + str(line_number)] = "CAS"
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS" + "\n")
                    # validación de si el numero de tarjeta está en blanco
                    if pd.isna(row['N_TARJETA']):
                        print("La fila " + str(line_number) + " no tiene número de tarjeta en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene número de tarjeta en " + file + "\n")
                        checkToReturn = False
                    # validación de si el codigo de promo esta en blanco o no es el especificado
                    if 'CADB1' != row['CODIGO_PROMO'] or pd.isna(row['CODIGO_PROMO']):
                        print("La fila " + str(
                            line_number) + " no tiene el codigo de promoción correcto en " + file + ". Se establece CADB1 por defecto")
                        sheet["Y" + str(line_number)] = 'CADB1'
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene el codigo de promoción correcto en " + file + ". Se establece CADB1 por defecto" + "\n")
                    # validación de la vía de impacto
                    if pd.isna(row['EMAIL']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL_SMS']) and pd.isna(row['TELEFONO_MOVIL_SMS_2']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS']), re.M) or re.fullmatch(
                                    "(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS_2']), re.M):
                                print(
                                    "La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                                with open('../OK.txt', 'a') as f:
                                    f.write("Fila " + str(line_number) + " su vía de impacto es SMS" "\n")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
                workbook.save(pathToFiles + file)
                workbook.close()
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn


class VentajasPlena:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Analizando " + file + " ....")
                print("Se debe agregar el target GRAL en " + file)
                workbook = load_workbook(filename=pathToFiles + file)
                sheet = workbook.active
                with open('../OK.txt', 'a') as f:
                    f.write("Se debe agregar el target GRAL en " + file + "\n")
                df = pd.read_excel(pathToFiles + file)
                line_number = 2
                for index, row in df.iterrows():
                    # validación del idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS")
                        sheet["C" + str(line_number)] = "CAS"
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS" + "\n")
                    # validación de si el numero de tarjeta está en blanco
                    if pd.isna(row['N_TARJETA']):
                        print("La fila " + str(line_number) + " no tiene número de tarjeta en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene número de tarjeta en " + file + "\n")
                        checkToReturn = False
                        # validación de si el codigo de promo esta en blanco o no es el especificado
                    if 'CAD2' != row['CODIGO_PROMO'] or pd.isna(row['CODIGO_PROMO']):
                        print("La fila " + str(
                            line_number) + " no tiene el codigo de promoción correcto en " + file + ". Se establece CAD2 por defecto")
                        sheet["Y" + str(line_number)] = 'CAD2'
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene el codigo de promoción correcto en " + file + ". Se establece CAD2 por defecto" + "\n")
                        # validación de la vía de impacto
                    if pd.isna(row['EMAIL']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL_SMS']) and pd.isna(row['TELEFONO_MOVIL_SMS_2']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS']), re.M) or re.fullmatch(
                                    "(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS_2']), re.M):
                                print(
                                    "La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
                workbook.save(pathToFiles + file)
                workbook.close()
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn


class RevisionMedica:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Analizando " + file + " ....")
                workbook = load_workbook(filename=pathToFiles + file)
                sheet = workbook.active
                df = pd.read_excel(pathToFiles + file)
                line_number = 2
                for index, row in df.iterrows():
                    # validación del idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS")
                        sheet["R" + str(line_number)] = "CAS"
                    # validación de si el numero de tarjeta está en blanco
                    if pd.isna(row['N_TARJETA']):
                        print("La fila " + str(line_number) + " no tiene número de tarjeta en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene número de tarjeta en " + file + "\n")
                        checkToReturn = False
                    # validación de si el codigo de promo esta en blanco o no es el especificado
                    if 'RMA' != row['CODIGO_PROMO'] or pd.isna(row['CODIGO_PROMO']):
                        print("La fila " + str(
                            line_number) + " no tiene el codigo de promoción correcto en " + file + ". Se establece RMA por defecto")
                        sheet["W" + str(line_number)] = 'RMA'
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene el codigo de promoción correcto en " + file + ". Se establece RMA por defecto" + "\n")
                    # validación de si no hay sexo especificado en las filas
                    if pd.isna(row['SEXO']):
                        print("La fila " + str(line_number) + " no tiene el sexo de la persona establecido en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene sexo especificado para la persona en" + file + "\n")
                        checkToReturn = False
                    # validacion de si no hay el tipo especificado en las filas
                    if 'TITULAR' != row['TIPO'] or 'BENEFICIARIO' != row['TIPO']:
                        print("La fila " + str(line_number) + " tiene un tipo distinto del establecido en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " tiene un tipo distinto del establecido en " + file + "\n")
                        checkToReturn = False
                    if pd.isna(row['EMAIL']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL_SMS']) and pd.isna(row['TELEFONO_MOVIL_SMS_2']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS']), re.M) or re.fullmatch(
                                    "(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS_2']), re.M):
                                print(
                                    "La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
                workbook.save(pathToFiles + file)
                workbook.close()
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn


# Se juntan la campaña 5 y 6 porque tienen las mismas validaciones
class SACAdeslasBasicaYPlena:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Se debe agregar el target GRAL en " + file)
                with open('../OK.txt', 'a') as f:
                    f.write("Se debe agregar el target GRAL en " + file + "\n")
                workbook = load_workbook(filename=pathToFiles + file)
                sheet = workbook.active
                df = pd.read_excel(pathToFiles + file)
                line_number = 2
                for index, row in df.iterrows():
                    # validación del idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + ". Se establece CAS por defecto")
                        sheet["C" + str(line_number)] = "CAS"
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS" + "\n")
                    if pd.isna(row['EMAIL']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL_SMS']) and pd.isna(row['TELEFONO_MOVIL_SMS_2']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS']), re.M) or re.fullmatch(
                                    "(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS_2']), re.M):
                                print(
                                    "La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
                workbook.save(pathToFiles + file)
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn


class CaixaAccidentes:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Analizando " + file + " ....")
                print("Se debe agregar el target GRAL en " + file)
                with open('../OK.txt', 'a') as f:
                    f.write("Se debe agregar el target GRAL en " + file + "\n")
                df = pd.read_excel(pathToFiles + file)
                workbook = load_workbook(filename=pathToFiles + file)
                sheet = workbook.active
                line_number = 2
                for index, row in df.iterrows():
                    # validación del idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + ". Se establece CAS por defecto")
                        sheet["C" + str(line_number)] = "CAS"
                        with open('../OK.txt', 'a') as f:
                            f.write("La fila " + str(
                                line_number) + " no tiene idioma en " + file + " .Se establece por defecto en CAS" + "\n")
                    # validacion si la columna de Marca está vacía
                    if pd.isna(row['MARCA']):
                        print("La fila " + str(line_number) + " no tiene datos en la columna MARCA en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene datos en la columna MARCA en " + file + "\n")
                        checkToReturn = False
                    if pd.isna(row['EMAIL']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL_SMS']) and pd.isna(row['TELEFONO_MOVIL_SMS_2']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS']), re.M) or re.fullmatch(
                                    "(666|696)+[ \d]?", str(row['TELEFONO_MOVIL_SMS_2']), re.M):
                                print(
                                    "La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
                workbook.save(pathToFiles + file)
                workbook.close()
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn


class PremiamosTuConfianzaDentalPost:
    excelfiles = []
    patternToExcel = ""

    def __init__(self, excelfiles, patternToExcel):
        self.excelfiles = excelfiles
        self.patternToExcel = patternToExcel

    def executeValidations(self):
        checkToReturn = True
        for file in self.excelfiles:
            if re.match(self.patternToExcel, file, re.M | re.I):
                print("Analizando " + file + " ....")
                df = pd.read_excel(pathToFiles + file)
                workbook = load_workbook(filename=pathToFiles + file)
                sheet = workbook.active
                line_number = 2
                for index, row in df.iterrows():
                    # validación de idioma
                    if pd.isna(row['IDIOMA']):
                        print("La fila " + str(
                            line_number) + " no tiene idioma en " + file + ". Se establece CAS por defecto")
                        sheet["C" + str(line_number)] = "CAS"
                    # validación de que venga NIF o número de tarjeta
                    if pd.isna(row['N_TARJETA']) and pd.isna(row['NIF']):
                        print("La fila " + str(line_number) + " no tiene número de tarjeta ni NIF en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene número de tarjeta ni NIF en " + file + "\n")
                        checkToReturn = False
                    if 'SRD4' != row['CODIGO_PROMO'] or pd.isna(row['CODIGO_PROMO']):
                        print("La fila " + str(line_number) + " no tiene el código promocional correcto en " + file)
                        with open('../validations.txt', 'a') as f:
                            f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                line_number) + " no tiene el código promocional correcto en " + file + "\n")
                        checkToReturn = False
                    if pd.isna(row['CORREO_CLIENTE']):
                        print("Fila " + str(line_number) + " no hay correo. Se revisan los telefonos")
                        if pd.isna(row['TELEFONO_MOVIL']):
                            print("La fila " + str(line_number) + " no tiene ni correo ni teléfono para ese cliente")
                            with open('../validations.txt', 'a') as f:
                                f.write(str(datetime.datetime.now()) + ": La fila " + str(
                                    line_number) + " no tiene ni correo ni teléfono para ese cliente en " + file + "\n")
                            checkToReturn = False
                        else:
                            if re.fullmatch("(666|696)+[ \d]?", str(row['TELEFONO_MOVIL']), re.M):
                                print("La fila " + str(line_number) + " no tiene números de teléfono válidos")
                                with open('../validations.txt', 'a') as f:
                                    f.write(str(datetime.datetime.now()) + ": En la fila " + str(
                                        line_number) + " los números de teléfono no son válidos en " + file + "\n")
                                checkToReturn = False
                            else:
                                print("Fila " + str(line_number) + " vía de impacto SMS")
                    else:
                        print("Fila " + str(line_number) + " vía de impacto CORREO")
                    line_number += 1
                workbook.save(pathToFiles + file)
                workbook.close()
            else:
                raise Exception(file + " is not an Excel File !")
        return checkToReturn
